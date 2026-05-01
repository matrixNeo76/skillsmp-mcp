"""
generate_xlsx.py — Genera skill_inventory.xlsx da skill_structure.json
con dati opzionali da SkillsMP (stelle, date, autore).

Uso:
  python generate_xlsx.py                          # XLSX base
  python generate_xlsx.py --with-skillsmp          # XLSX con dati SkillsMP
"""
import sys, os, json, time, argparse
sys.stdout.reconfigure(encoding='utf-8')
os.environ['PYTHONIOENCODING'] = 'utf-8'

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.chart import BarChart, Reference

BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
STRUCTURE_PATH = os.path.join(BASE_DIR, 'data', 'skill_structure.json')
OUTPUT_PATH = os.path.join(BASE_DIR, 'docs', 'skill_inventory.xlsx')

# Stili
HEADER_FONT = Font(name='Calibri', bold=True, color='FFFFFF', size=11)
HEADER_FILL = PatternFill(start_color='2F5496', end_color='2F5496', fill_type='solid')
DOMAIN_FONT = Font(name='Calibri', bold=True, size=12, color='2F5496')
SUBDOMAIN_FONT = Font(name='Calibri', bold=True, size=10, color='5B9BD5')
SUBDOMAIN_FILL = PatternFill(start_color='D6E4F0', end_color='D6E4F0', fill_type='solid')
STAR_FILL = PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid')  # gold for high stars
THIN_BORDER = Border(
    left=Side(style='thin', color='D9D9D9'),
    right=Side(style='thin', color='D9D9D9'),
    top=Side(style='thin', color='D9D9D9'),
    bottom=Side(style='thin', color='D9D9D9')
)

# ── Load structure ───────────────────────────────────────────
with open(STRUCTURE_PATH, 'r', encoding='utf-8') as f:
    structure = json.load(f)

HEADERS = ['#', 'Nome Skill', 'Dominio', 'Sottodominio', 'Cosa Fa',
           '⭐ Stelle', '📅 Aggiornamento', '👤 Autore', 'Da Tenere?', 'Note']
COL_WIDTHS = [5, 42, 22, 20, 50, 12, 16, 22, 12, 30]

# ── SkillsMP check opzionale ─────────────────────────────────
def fetch_skillsmp(skill_name):
    """Cerca skill su SkillsMP, restituisce (stars, updated, author)."""
    api_key = os.environ.get('SKILLSMP_API_KEY', '')
    if not api_key:
        return (0, '-', '-')
    try:
        import httpx
        r = httpx.get(
            'https://skillsmp.com/api/v1/skills/search',
            params={'q': skill_name.replace('-', ' '), 'limit': 1, 'sortBy': 'stars'},
            headers={'Authorization': f'Bearer {api_key}'},
            timeout=10
        )
        data = r.json()
        skills = data.get('data', {}).get('skills', [])
        if skills:
            s = skills[0]
            stars = s.get('stars', 0)
            updated = s.get('updatedAt', '')
            if updated:
                updated = time.strftime('%Y-%m-%d', time.gmtime(int(updated)))
            author = s.get('author', '')
            return (int(stars), updated, author)
    except:
        pass
    return (0, '-', '-')

# ── Descrizioni (da file locale se possibile) ─────────────────
def get_skill_description(skill_name):
    """Legge descrizione dal SKILL.md locale se esiste."""
    skill_path = os.path.expanduser(f'~/.agents/skills/{skill_name}/SKILL.md')
    if os.path.exists(skill_path):
        try:
            with open(skill_path, 'r', encoding='utf-8') as f:
                content = f.read()
            # Estrai frontmatter YAML
            import re
            m = re.search(r'description:\s*"([^"]*)"', content)
            if m:
                return m.group(1)[:120]
            m = re.search(r'description:\s*([^\n]+)', content)
            if m:
                return m.group(1).strip()[:120]
        except:
            pass
    return ''

# ── Generazione XLSX ─────────────────────────────────────────
def generate(with_skillsmp=False):
    wb = Workbook()
    ws = wb.active
    ws.title = 'Skill Inventory'

    # Header
    for col, (h, w) in enumerate(zip(HEADERS, COL_WIDTHS), 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        ws.column_dimensions[get_column_letter(col)].width = w

    # Dropdown per 'Da Tenere?'
    dv = DataValidation(type='list', formula1='"Si,No,Da valutare,Archiviare"', allow_blank=True)
    dv.error = 'Scegli: Si, No, Da valutare, Archiviare'
    dv.errorTitle = 'Valore non valido'
    ws.add_data_validation(dv)

    row = 2
    skill_counter = 0
    total_with_stars = 0

    for domain in structure['domains']:
        dn = domain['name']
        dnum = domain['number']

        # Riga dominio
        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=len(HEADERS))
        cell = ws.cell(row=row, column=2, value=f'{dnum}. {dn}')
        cell.font = DOMAIN_FONT
        row += 1

        for sub in domain['subdomains']:
            sub_name = sub['name']
            skills = sub['skills']

            # Riga sottodominio
            ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=len(HEADERS))
            cell = ws.cell(row=row, column=2, value=f'  {sub_name}')
            cell.font = SUBDOMAIN_FONT
            cell.fill = SUBDOMAIN_FILL
            row += 1

            for skill_name in skills:
                skill_counter += 1
                desc = get_skill_description(skill_name)

                ws.cell(row=row, column=1, value=skill_counter).border = THIN_BORDER
                ws.cell(row=row, column=2, value=skill_name).border = THIN_BORDER
                ws.cell(row=row, column=3, value=dn).border = THIN_BORDER
                ws.cell(row=row, column=4, value=sub_name).border = THIN_BORDER
                ws.cell(row=row, column=5, value=desc).border = THIN_BORDER

                # SkillsMP data
                if with_skillsmp:
                    stars, updated, author = fetch_skillsmp(skill_name)
                    if stars > 0:
                        total_with_stars += 1
                    ws.cell(row=row, column=6, value=stars).border = THIN_BORDER
                    ws.cell(row=row, column=6).alignment = Alignment(horizontal='right')
                    if stars >= 10000:
                        ws.cell(row=row, column=6).fill = STAR_FILL
                    ws.cell(row=row, column=7, value=updated).border = THIN_BORDER
                    ws.cell(row=row, column=8, value=author).border = THIN_BORDER
                else:
                    for c in [6, 7, 8]:
                        ws.cell(row=row, column=c, value='').border = THIN_BORDER

                # Da Tenere?
                cell_keep = ws.cell(row=row, column=9, value='')
                cell_keep.border = THIN_BORDER
                cell_keep.alignment = Alignment(horizontal='center')
                dv.add(cell_keep)

                ws.cell(row=row, column=10, value='').border = THIN_BORDER

                # Alterna colori
                if skill_counter % 2 == 0:
                    for c in range(1, len(HEADERS) + 1):
                        ws.cell(row=row, column=c).fill = PatternFill(
                            start_color='F5F5F5', end_color='F5F5F5', fill_type='solid')

                row += 1

        row += 1  # spazio tra domini

    # Freeze panes + auto filter
    ws.freeze_panes = 'A2'
    ws.auto_filter.ref = f'A1:{get_column_letter(len(HEADERS))}{row - 1}'

    # ── Foglio Statistiche ──
    ws2 = wb.create_sheet('Statistiche')
    ws2.cell(row=1, column=1, value='Dominio').font = HEADER_FONT
    ws2.cell(row=1, column=1).fill = HEADER_FILL
    ws2.cell(row=2, column=1, value='Tutte le skill').font = HEADER_FONT
    ws2.cell(row=2, column=1).fill = HEADER_FILL

    for col, h in enumerate(['Skill', '⭐ Media Stelle', '⭐ Max Stelle', 'Da Tenere (Si)', 'Da Tenere (No)', 'Da Valutare', 'Archiviare'], 2):
        cell = ws2.cell(row=1, column=col, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell = ws2.cell(row=2, column=col, value=skill_counter)
        cell.font = Font(bold=True)

    ws2.column_dimensions['A'].width = 30
    for c in range(2, 8):
        ws2.column_dimensions[get_column_letter(c)].width = 18

    r = 3
    for domain in structure['domains']:
        dn = domain['name']
        dnum = domain['number']
        total = sum(len(sub['skills']) for sub in domain['subdomains'])
        ws2.cell(row=r, column=1, value=f'{dnum}. {dn}').font = Font(bold=True)
        ws2.cell(row=r, column=2, value=total)
        r += 1

    wb.save(OUTPUT_PATH)
    return skill_counter, total_with_stars

# ── Main ──
if __name__ == '__main__':
    parser = argparse.ArgumentParser(description='Genera skill_inventory.xlsx')
    parser.add_argument('--with-skillsmp', action='store_true',
                       help='Arricchisce con dati SkillsMP (stelle, date, autore)')
    parser.add_argument('--output', default=OUTPUT_PATH, help='Percorso output')
    args = parser.parse_args()

    OUTPUT_PATH = args.output
    total, stars = generate(with_skillsmp=args.with_skillsmp)
    print(f'XLSX generato: {OUTPUT_PATH}')
    print(f'{total} skill in {len(structure["domains"])} domini')
    print(f'{os.path.getsize(OUTPUT_PATH)} bytes')
    if args.with_skillsmp:
        print(f'{stars} skill verificate su SkillsMP')
    print(f'\n  Prossimo passo: python generate_xlsx.py --with-skillsmp')
    print(f'  per arricchire con stelle SkillsMP (richiede API key)')
